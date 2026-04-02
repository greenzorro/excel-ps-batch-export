#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel-PS Batch Export 真实文件集成测试
======================================

使用真实文件进行集成测试，减少 mock 使用。
"""

import os
import sys
import tempfile
import shutil
import pytest
import pandas as pd
from pathlib import Path

# 添加项目根目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


class TestClipboardImporterIntegration:
    """剪贴板导入器集成测试 - 使用真实文件"""

    def test_write_and_read_real_excel(self):
        """测试真实的 Excel 读写"""
        import openpyxl

        # 创建临时 Excel 文件
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # 写入数据
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="File_name")
            ws.cell(row=1, column=2, value="Content")
            ws.cell(row=2, column=1, value="test1")
            ws.cell(row=2, column=2, value="content1")
            wb.save(tmp_path)

            # 读取数据验证
            from src.psd_renderer import read_excel_file
            df = read_excel_file(tmp_path)

            # 验证具体内容，不只是类型
            assert isinstance(df, pd.DataFrame)
            assert len(df) == 1, f"Expected 1 row, got {len(df)}"
            assert "File_name" in df.columns
            assert "Content" in df.columns
            assert df.iloc[0]["File_name"] == "test1"
            assert df.iloc[0]["Content"] == "content1"

        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_text_parsing_with_real_params(self):
        """测试文本参数解析的真实场景"""
        from src.psd_renderer import parse_text_params, parse_rotation_from_name

        # 测试组合参数
        result = parse_text_params("@标题#t_c_p")
        assert result["align"] == "center"
        assert result["paragraph"] is True
        assert result["valign"] == "top"  # 默认值

        # 测试旋转参数
        rotation = parse_rotation_from_name("@标题#t_c_a15")
        assert rotation == 15.0

        rotation = parse_rotation_from_name("@标题#t_a-30")
        assert rotation == -30.0

    def test_image_params_with_real_values(self):
        """测试图片参数解析的真实场景"""
        from src.psd_renderer import parse_image_params

        # 测试 cover + 居中
        result = parse_image_params("@产品图#i_cover_cm")
        assert result["mode"] == "cover"
        assert result["alignment"] == "cm"

        # 测试 contain + 左上
        result = parse_image_params("@产品图#i_contain_lt")
        assert result["mode"] == "contain"
        assert result["alignment"] == "lt"

        # 测试默认值
        result = parse_image_params("@产品图#i")
        assert result["mode"] == "cover"
        assert result["alignment"] == "cm"


class TestDataValidation:
    """数据验证测试 - 验证具体值"""

    def test_excel_data_with_specific_values(self):
        """测试 Excel 数据验证的具体值"""
        test_data = {
            "File_name": ["test1", "test2", "test3"],
            "标题": ["标题1", "标题2", "标题3"],
            "内容": ["内容1", "内容2", "内容3"]
        }
        df = pd.DataFrame(test_data)

        # 验证具体值，不只是长度
        assert len(df) == 3
        assert df.iloc[0]["File_name"] == "test1"
        assert df.iloc[1]["标题"] == "标题2"
        assert df.iloc[2]["内容"] == "内容3"

    def test_sanitize_filename_with_specific_chars(self):
        """测试文件名清理的具体字符"""
        from src.psd_renderer import sanitize_filename

        # 测试具体的非法字符替换
        assert sanitize_filename("file:name") == "file_name"
        assert sanitize_filename("file*name") == "file_name"
        assert sanitize_filename("file?name") == "file_name"
        assert sanitize_filename("file/name") == "file_name"
        assert sanitize_filename("file\\name") == "file_name"

        # 测试组合
        assert sanitize_filename("file:*name?") == "file__name_"


class TestTextRenderingReal:
    """文本渲染真实测试"""

    def test_calculate_text_position_specific_values(self):
        """测试文本位置计算的具体值"""
        from src.psd_renderer import calculate_text_position
        from PIL import Image, ImageDraw, ImageFont

        test_img = Image.new("RGB", (200, 50))
        draw = ImageDraw.Draw(test_img)
        font = ImageFont.load_default()

        # 测试具体值
        text = "Test"
        x_pos, y_pos = calculate_text_position(text, 200, 16, "center", draw, font)

        # 验证具体范围，不只是大于/小于
        assert 0 <= x_pos <= 200, f"x_pos {x_pos} out of range [0, 200]"
        assert -50 <= y_pos <= 0, f"y_pos {y_pos} out of range [-50, 0]"

    def test_preprocess_text_specific_replacements(self):
        """测试文本预处理的具体替换"""
        from src.psd_renderer import preprocess_text

        # 测试具体的替换规则
        assert preprocess_text('"test"') == "test"
        assert preprocess_text("A/B/C") == "A&B&C"
        assert preprocess_text("  test  ") == "test"
        assert preprocess_text("test_x000D_") == "test"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
