'''
File: clipboard_importer.py
Project: excel-ps-batch-export
Created: 2025-10-09
Author: Victor Cheng
Email: hi@victor42.work
Description:
剪贴板导入器 - 从系统剪贴板读取表格数据并写入Excel文件，自动触发图片渲染

主要功能：
1. 从系统剪贴板读取表格数据（支持制表符和逗号分隔格式）
2. 自动解析剪贴板数据为DataFrame格式
3. 在当前目录查找Excel文件（支持多文件选择）
4. 确定目标sheet：优先使用名为"粘贴"的sheet，否则使用第二个sheet
5. 清空从B2单元格开始的右下方所有区域
6. 从B2单元格开始写入解析后的数据
7. 支持文件选择退出功能（输入"q"退出或Ctrl+C中断）
8. 自动运行PSD渲染器生成图片，实现数据导入到图片生成的一体化流程

使用场景：
- 快速将网页表格、其他软件表格数据导入到Excel中
- 避免手动打开Excel复制粘贴的繁琐操作
- 实现"数据导入→图片生成"的一键式工作流程

运行方式：
python clipboard_importer.py
'''

import sys
import os
import pandas as pd
import pyperclip
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import xlwings as xw
import subprocess

# ===== 配置项 =====
# 渲染配置
DEFAULT_FORMAT = 'jpg'      # 默认输出格式
RENDER_TIMEOUT = 300        # 渲染超时时间（秒）

# 文件路径配置
EXPORT_DIR = "export"       # 输出目录

def safe_print_message(message):
    """安全打印消息，处理Windows控制台编码问题
    
    :param str message: 要打印的消息
    """
    try:
        print(message)
    except UnicodeEncodeError:
        # 如果直接打印失败，使用安全的编码方式
        safe_message = message.encode('ascii', errors='replace').decode('ascii')
        print(safe_message)

def get_clipboard_data():
    """从系统剪贴板读取数据
    
    :return str: 剪贴板内容
    :raises Exception: 当无法读取剪贴板时抛出异常
    """
    try:
        clipboard_content = pyperclip.paste()
        if not clipboard_content or clipboard_content.strip() == '':
            raise ValueError("剪贴板为空或只包含空白字符")
        return clipboard_content
    except Exception as e:
        raise Exception(f"无法读取剪贴板: {str(e)}")

def parse_clipboard_data(clipboard_content):
    """解析剪贴板数据为DataFrame

    :param str clipboard_content: 剪贴板内容
    :return pd.DataFrame: 解析后的数据
    """
    # 尝试按制表符分割（Excel复制格式）
    lines = clipboard_content.strip().split('\n')

    # 检查是否为空数据（只有空字符串或空白字符）
    if not lines or (len(lines) == 1 and not lines[0].strip()):
        raise ValueError("剪贴板数据为空")
    
    # 解析每行数据
    data = []
    for line in lines:
        # 按制表符分割，如果制表符不存在则按逗号分割
        if '\t' in line:
            row_data = line.split('\t')
        elif ',' in line:
            row_data = line.split(',')
        else:
            # 如果没有分隔符，将整行作为一个单元格
            row_data = [line]

        # 注意：不再清理空白字符，统一在PSD渲染阶段处理
        data.append(row_data)
    
    # 转换为DataFrame
    if len(data) == 1:
        # 只有一行数据，作为单行处理
        df = pd.DataFrame([data[0]])
    else:
        # 多行数据，所有行都作为数据，不把第一行作为列名
        df = pd.DataFrame(data)
    
    return df

def find_target_excel_file():
    """在当前目录查找Excel文件

    :return str: Excel文件路径
    :raises FileNotFoundError: 当找不到Excel文件时抛出异常
    """
    excel_files = [f for f in os.listdir('.') if f.endswith(('.xlsx', '.xls'))]

    # 按文件名排序
    excel_files.sort()

    if not excel_files:
        raise FileNotFoundError("当前目录未找到Excel文件")

    # 如果有多个Excel文件，让用户选择
    if len(excel_files) > 1:
        safe_print_message("找到多个Excel文件:")
        for i, file in enumerate(excel_files, 1):
            safe_print_message(f"  {i}. {file}")

        try:
            choice = input("请选择文件编号 (默认1，输入q退出): ").strip()

            # 检查是否要退出
            if choice.lower() in ['q', 'quit', 'exit', '退出']:
                safe_print_message("用户选择退出程序")
                sys.exit(0)

            if not choice:
                choice = 1
            else:
                choice = int(choice)

            if 1 <= choice <= len(excel_files):
                return excel_files[choice - 1]
            else:
                raise ValueError("无效的选择")
        except ValueError:
            safe_print_message("无效选择，使用第一个文件")
            return excel_files[0]
        except KeyboardInterrupt:
            safe_print_message("\n用户中断程序")
            sys.exit(1)
        except EOFError:
            # 处理非交互式环境（如测试环境）
            safe_print_message("非交互式环境，使用第一个文件")
            return excel_files[0]

    return excel_files[0]

def get_target_sheet(workbook):
    """获取目标sheet名称
    
    :param Workbook workbook: openpyxl Workbook对象
    :return str: 目标sheet名称
    """
    sheet_names = workbook.sheetnames
    
    # 优先查找名为"粘贴"的sheet
    if "粘贴" in sheet_names:
        return "粘贴"
    
    # 如果没有"粘贴"sheet，使用第二个sheet
    if len(sheet_names) >= 2:
        return sheet_names[1]  # 第二个sheet
    
    # 如果只有一个sheet，使用第一个sheet
    return sheet_names[0]

def write_to_excel(excel_file, df):
    """将DataFrame写入Excel文件的指定sheet，从B2单元格开始

    :param str excel_file: Excel文件路径
    :param pd.DataFrame df: 要写入的数据
    :raises Exception: 当写入失败时抛出异常
    """
    try:
        # 加载现有工作簿
        workbook = load_workbook(excel_file)
        target_sheet_name = get_target_sheet(workbook)

        # 获取或创建目标sheet
        if target_sheet_name in workbook.sheetnames:
            sheet = workbook[target_sheet_name]
        else:
            sheet = workbook.create_sheet(target_sheet_name)

        # 清空从B2开始的右下方所有区域
        max_row = sheet.max_row
        max_col = sheet.max_column

        # 计算需要清空的范围：从第2行到最大行，从第2列到最大列
        clear_start_row = 2
        clear_end_row = max(max_row, 100)  # 至少清空到第100行
        clear_start_col = 2
        clear_end_col = max(max_col, 26)   # 至少清空到Z列

        safe_print_message(f"正在清空 {target_sheet_name} 的 B2 开始区域 (行{clear_start_row}-{clear_end_row}, 列{clear_start_col}-{clear_end_col})...")

        for row in range(clear_start_row, clear_end_row + 1):
            for col in range(clear_start_col, clear_end_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.value = None

        # 从B2单元格开始写入数据
        start_row = 2  # B2单元格的行号
        start_col = 2  # B2单元格的列号

        # 写入数据
        for r_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start_row):
            for c_idx, value in enumerate(row_data, start_col):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # 保存文件
        workbook.save(excel_file)

        # 强制Excel重新计算公式（使用xlwings确保公式正确更新）
        try:
            safe_print_message("正在重新计算公式...")

            # 使用xlwings打开Excel并重新计算公式
            app = xw.App(visible=False)
            wb = app.books.open(excel_file)
            wb.app.calculate()
            wb.save()
            wb.close()
            app.quit()
            safe_print_message("公式重新计算完成")

        except Exception as e:
            safe_print_message(f"警告: 无法重新计算公式: {str(e)}")
            safe_print_message("第一个sheet的数据可能需要手动刷新")

        return target_sheet_name, start_row, len(df)

    except Exception as e:
        raise Exception(f"写入Excel文件失败: {str(e)}")

def get_matching_psds(excel_file):
    """获取匹配的PSD文件列表
    
    :param str excel_file: Excel文件名
    :return list: 匹配的PSD文件列表
    """
    base_name = os.path.splitext(excel_file)[0]
    matching_psds = []
    for f in os.listdir('.'):
        if f.endswith('.psd'):
            # 提取文件名前缀（第一个井号前的部分）
            name_without_ext = os.path.splitext(f)[0]
            if '#' in name_without_ext:
                prefix = name_without_ext.split('#', 1)[0]
            else:
                prefix = name_without_ext
            if prefix == base_name:
                matching_psds.append(f)
    return matching_psds

def run_psd_renderer(excel_file):
    """运行PSD渲染器进行自动图片导出

    :param str excel_file: Excel文件路径
    :return bool: 渲染是否成功
    """
    # 从Excel文件名提取模板前缀（去掉.xlsx后缀）
    template_name = os.path.splitext(excel_file)[0]

    safe_print_message(f"\n正在准备自动渲染图片...")
    safe_print_message(f"Excel文件: {excel_file}")

    # 查找匹配的PSD模板
    matching_psds = get_matching_psds(excel_file)

    if not matching_psds:
        safe_print_message(f"警告: 未找到与 '{template_name}' 匹配的PSD模板文件")
        safe_print_message("请确保PSD文件命名格式为: [Excel前缀]#[后缀].psd")
        return False

    safe_print_message(f"找到 {len(matching_psds)} 个匹配的PSD模板:")
    for i, psd_file in enumerate(matching_psds, 1):
        safe_print_message(f"  {i}. {psd_file}")

    safe_print_message("正在自动启动PSD渲染器...")

    # 构建命令行参数
    cmd = [
        sys.executable,  # 使用当前Python解释器
        "psd_renderer.py",
        template_name,
        DEFAULT_FORMAT
    ]

    # 运行PSD渲染器（psd_renderer.py内部已处理所有异常）
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=RENDER_TIMEOUT)

    if result.returncode == 0:
        safe_print_message("\n✓ 图片渲染成功!")
        safe_print_message(f"输出目录: {EXPORT_DIR}/")
        safe_print_message(f"输出格式: {DEFAULT_FORMAT}")
        safe_print_message(f"处理的PSD模板: {len(matching_psds)} 个")
        return True
    else:
        safe_print_message(f"\n✗ 图片渲染失败:")
        if result.stdout:
            safe_print_message(f"输出: {result.stdout}")
        if result.stderr:
            safe_print_message(f"错误: {result.stderr}")
        return False

def main():
    """主函数"""
    try:
        # 切换到脚本所在目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
        os.chdir(script_dir)

        safe_print_message("剪贴板导入器启动...")

        # 1. 读取剪贴板数据
        safe_print_message("正在读取剪贴板数据...")
        clipboard_content = get_clipboard_data()
        safe_print_message(f"读取到 {len(clipboard_content)} 个字符")

        # 2. 解析剪贴板数据
        safe_print_message("正在解析剪贴板数据...")
        df = parse_clipboard_data(clipboard_content)
        safe_print_message(f"解析为 {len(df)} 行 {len(df.columns)} 列数据")

        # 3. 查找目标Excel文件
        safe_print_message("正在查找Excel文件...")
        excel_file = find_target_excel_file()
        safe_print_message(f"目标文件: {excel_file}")

        # 4. 写入Excel文件
        safe_print_message("正在写入Excel文件...")
        sheet_name, start_row, row_count = write_to_excel(excel_file, df)

        # 5. 输出结果
        safe_print_message(f"\n✓ 导入成功!")
        safe_print_message(f"   文件: {excel_file}")
        safe_print_message(f"   Sheet: {sheet_name}")
        safe_print_message(f"   位置: 第{start_row}行开始")
        safe_print_message(f"   数据: {row_count}行 {len(df.columns)}列")

        # 6. 自动运行PSD渲染器
        run_psd_renderer(excel_file)

        return 0

    except Exception as e:
        safe_print_message(f"\n✗ 错误: {str(e)}")
        safe_print_message("\n使用提示:")
        safe_print_message("  1. 请确保剪贴板中有表格数据")
        safe_print_message("  2. 请确保当前目录有Excel文件")
        safe_print_message("  3. 如果Excel文件已打开，请先关闭或确保未锁定")
        return 1

if __name__ == "__main__":
    sys.exit(main())

